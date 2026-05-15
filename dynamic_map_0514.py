"""
EU Installer 客户地理分布 – 动态交互 HTML 地图生成器 v7
Sheet1 = Max Installer   (header第12行, 数据第14行起)
Sheet2 = Thomas Installers (header第11行, 数据第13行起)
依赖: pip install openpyxl requests
输出: eu_customers_map.html
"""

import re
import time
import json
import requests
import openpyxl
from pathlib import Path

# ─── 1. 配置 ──────────────────────────────────────────────────
DATA_FILE   = "D:/Ethan.He/Desktop/eu_installer.xlsx"   # ← 修改为你的路径
OUTPUT_HTML = "eu_customers_map.html"

# ─── 2. 国家识别 ──────────────────────────────────────────────
REGION_TO_COUNTRY = {
    "DE": "Germany", "de": "Germany",
    "AT": "Austria", "at": "Austria",
    "CH": "Switzerland", "ch": "Switzerland",
    "FR": "France",  "NL": "Netherlands", "BE": "Belgium",
    "PL": "Poland",  "IT": "Italy", "ES": "Spain",
}
COUNTRY_HINTS = {
    "Deutschland": "Germany", "Österreich": "Austria",
    "Schweiz": "Switzerland", "Switzerland": "Switzerland",
    "Austria": "Austria", "Germany": "Germany",
}
AUSTRIA_REGIONS = {
    "Steiermark","Tirol","Vorarlberg","Niederösterreich",
    "Oberösterreich","Salzburg","Wien","Kärnten","Burgenland",
}

def region_to_country(region_code: str) -> str:
    if not region_code:
        return "Germany"
    return REGION_TO_COUNTRY.get(str(region_code).strip(), "Germany")

def detect_country_from_text(text: str) -> str:
    for kw, country in COUNTRY_HINTS.items():
        if kw in text:
            return country
    for region in AUSTRIA_REGIONS:
        if region in text:
            return "Austria"
    return "Germany"

def extract_city_from_address(address: str) -> str:
    if not address:
        return ""
    address = re.sub(r"[\t]", " ", str(address)).strip()
    lines = [l.strip() for l in re.split(r"[\n,，]", address) if l.strip()]
    for line in reversed(lines):
        clean = re.sub(r"[\u202f\u2003\u00a0]", " ", line).strip()
        if clean in COUNTRY_HINTS or clean in AUSTRIA_REGIONS:
            continue
        if re.match(r"^\d{4,5}\s+", clean):
            part = re.sub(r"^\d{4,5}\s+", "", clean)
            if part:
                return part
        elif re.match(r"^\d+$", clean):
            continue
        else:
            return clean
    return lines[-1] if lines else address.strip()

def clean_str(val) -> str:
    if val is None:
        return ""
    return re.sub(r"[\t]", " ", str(val)).replace("\n", " ").strip()

# ─── 3. 级别分类 ──────────────────────────────────────────────
def classify(level: str) -> str:
    l = level.lower()
    if "wholesaler" in l:
        return "Wholesaler"
    elif "large" in l or "big" in l:
        return "Large"
    elif "medium" in l:
        return "Medium"
    else:
        return "Small"

# ─── 4. Sheet 配置 ────────────────────────────────────────────
# 新文件结构:
#
# Sheet1: "Max Installer"
#   header 第12行, 数据从第14行
#   col1=Act, col2=level, col3=Status, col7=Region, col8=Inverter
#   col10=Company, col14=Contact Person, col16=Phone
#   (无City/Address列 → 用公司名+Region geocode)
#
# Sheet2: "Thomas Installers"
#   header 第11行, 数据从第13行
#   col0=Act, col1=Level, col2=Status, col6=Region, col7=Inverter
#   col10=City, col12=Company, col16=Contact Person
#   col19=Phone, col21=Address, col9=Country(de/at/ch)

wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
sheet_names = wb.sheetnames
print(f"发现 Sheets: {sheet_names}")

SHEET_CONFIG = {
    sheet_names[0]: {   # Max Installer
        "owner":      "Max",
        "data_start": 14,
        "col": {
            "act": 1, "level": 2, "status": 3,
            "region": 7, "inverter": 8,
            "company": 10, "contact": 14, "phone": 16,
            "city": None, "address": None, "country_col": None,
        }
    },
    sheet_names[1]: {   # Thomas Installers
        "owner":      "Thomas",
        "data_start": 13,
        "col": {
            "act": 0, "level": 1, "status": 2,
            "region": 6, "inverter": 7,
            "company": 12, "contact": 16, "phone": 19,
            "city": 10, "address": 21, "country_col": 9,
        }
    },
}

def safe_get(row, idx):
    if idx is None:
        return None
    return row[idx] if idx < len(row) else None

rows = []
for sheet_name, cfg in SHEET_CONFIG.items():
    ws    = wb[sheet_name]
    owner = cfg["owner"]
    c     = cfg["col"]
    count = 0

    for row in ws.iter_rows(min_row=cfg["data_start"], values_only=True):
        company = clean_str(safe_get(row, c["company"]))
        if not company or company.lower() in ("none", "—", ""):
            continue

        level    = clean_str(safe_get(row, c["level"]))
        status   = clean_str(safe_get(row, c["status"]))
        act      = clean_str(safe_get(row, c["act"]))
        region   = clean_str(safe_get(row, c["region"]))
        inverter = clean_str(safe_get(row, c["inverter"]))
        contact  = clean_str(safe_get(row, c["contact"]))
        phone    = clean_str(safe_get(row, c["phone"]))

        city_raw    = clean_str(safe_get(row, c["city"]))
        address_raw = clean_str(safe_get(row, c["address"]))
        country_raw = clean_str(safe_get(row, c["country_col"]))

        # 城市：优先city列，其次从address提取
        city = city_raw if city_raw else extract_city_from_address(address_raw)

        # 国家：优先专用country列，其次region，其次从address文本推断
        if country_raw:
            country = region_to_country(country_raw)
        elif region:
            country = region_to_country(region)
        else:
            country = detect_country_from_text(address_raw + " " + company)

        rows.append({
            "city":     city,
            "company":  company,
            "level":    level or "Unknown",
            "status":   status or "Unknown",
            "act":      act or "",
            "inverter": inverter or "—",
            "contact":  contact or "—",
            "phone":    phone or "—",
            "country":  country,
            "owner":    owner,
        })
        count += 1

    print(f"  [{sheet_name}] ({owner}): 读取 {count} 条")

print(f"\n✅ 合计读取 {len(rows)} 条记录")

# ─── 5. 地理编码（通用三级策略）─────────────────────────────────
COUNTRY_DEFAULTS = {
    "Germany":     (51.1657,  10.4515),
    "Austria":     (47.8095,  13.0550),
    "Switzerland": (46.8182,   8.2275),
    "France":      (46.2276,   2.2137),
    "Netherlands": (52.1326,   5.2913),
    "Belgium":     (50.5039,   4.4699),
    "Poland":      (51.9194,  19.1451),
    "Italy":       (41.8719,  12.5674),
    "Spain":       (40.4168,  -3.7038),
}

def _nominatim(params: dict):
    url = "https://nominatim.openstreetmap.org/search"
    headers = {"User-Agent": "EU-InstallerMap/7.0"}
    try:
        r = requests.get(url, params={**params, "format": "json", "limit": 1},
                         headers=headers, timeout=10)
        data = r.json()
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception:
        pass
    return None, None

def geocode_universal(city: str, company: str, country: str, has_address: bool):
    """
    三级兜底：
    1. 有城市 → 城市名搜索（精确）
    2. 无城市 → 公司名+国家搜索（近似）
    3. 均失败 → 国家中心坐标（国家级）
    """
    if has_address and city:
        lat, lon = _nominatim({"city": city, "country": country})
        if lat:
            return lat, lon, "exact"
        time.sleep(0.5)

    lat, lon = _nominatim({"q": f"{company}, {country}"})
    if lat:
        return lat, lon, "approximate"
    time.sleep(0.5)

    default = COUNTRY_DEFAULTS.get(country, (51.1657, 10.4515))
    return default[0], default[1], "country"

print("正在获取坐标（三级兜底策略，每次间隔1秒）...")
records = []
seen = {}
for item in rows:
    has_address = bool(item["city"] and item["city"] != item["company"])
    key = (item["city"], item["company"], item["country"])
    if key in seen:
        lat, lon, precision = seen[key]
    else:
        lat, lon, precision = geocode_universal(
            item["city"], item["company"], item["country"], has_address
        )
        seen[key] = (lat, lon, precision)
        time.sleep(1)

    icon = {"exact": "✓", "approximate": "~", "country": "⚠"}[precision]
    records.append({
        **item,
        "category":  classify(item["level"]),
        "precision": precision,
        "lat": lat, "lon": lon,
    })
    print(f"  {icon} [{item['owner']}][{item['country']}] {item['city'] or '(no city)'} | {item['company']} [{precision}]")

# ─── 6. 统计 ──────────────────────────────────────────────────
summary = {"total": len(records), "by_owner": {}, "by_category": {}, "by_status": {}, "by_country": {}}
for r in records:
    for k, f in [("by_owner","owner"),("by_category","category"),("by_status","status"),("by_country","country")]:
        summary[k][r[f]] = summary[k].get(r[f], 0) + 1

# ─── 7. GeoJSON ───────────────────────────────────────────────
features = []
for r in records:
    features.append({
        "type": "Feature",
        "properties": {k: r[k] for k in
            ["city","company","level","category","status","act","inverter",
             "contact","phone","country","owner","precision"]},
        "geometry": {"type": "Point", "coordinates": [r["lon"], r["lat"]]},
    })

geojson_data = json.dumps({"type": "FeatureCollection", "features": features}, ensure_ascii=False)
summary_json = json.dumps(summary, ensure_ascii=False)

# ─── 8. HTML ──────────────────────────────────────────────────
all_countries = sorted(summary["by_country"].keys())
country_pills_html = "\n    ".join(
    f'<div class="pill active" data-country="{c}" onclick="toggleCountry(\'{c}\')">'
    f'{REGION_TO_COUNTRY.get(c, c)[:2].upper() if len(c) > 2 else c}</div>'
    for c in all_countries
)

HTML = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EU Installer Map</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#f5f0e8;--panel:#fffdf8;--panel2:#f0ebe0;
  --border:#d8cfc0;--text:#2c2416;--muted:#8a7a65;
  --accent:#2563eb;--green:#16a34a;--orange:#ea580c;
  --radius:10px;--shadow:0 2px 16px rgba(0,0,0,.10);
  --font:'Segoe UI',system-ui,sans-serif;
  --thomas:#0ea5e9;--max:#f97316;
}}
body{{font-family:var(--font);background:var(--bg);color:var(--text);height:100vh;display:flex;flex-direction:column;overflow:hidden}}

header{{
  padding:8px 14px;background:var(--panel);border-bottom:1.5px solid var(--border);
  display:flex;align-items:center;gap:10px;flex-wrap:wrap;z-index:1000;box-shadow:var(--shadow)
}}
header h1{{font-size:14px;font-weight:700;white-space:nowrap;color:var(--text)}}
header h1 span{{color:var(--accent)}}

.pill-group{{display:flex;gap:5px;flex-wrap:wrap;align-items:center}}
.sep{{width:1px;height:18px;background:var(--border);margin:0 3px;flex-shrink:0}}
.group-label{{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}}
.pill{{
  display:flex;align-items:center;gap:4px;padding:3px 10px;
  border-radius:20px;border:1.5px solid var(--border);background:transparent;
  color:var(--muted);font-size:11px;cursor:pointer;transition:all .15s;user-select:none
}}
.pill .dot{{width:7px;height:7px;border-radius:50%;flex-shrink:0}}
.pill.active{{color:var(--text);background:rgba(0,0,0,.05)}}
.pill[data-cat="Large"].active{{border-color:#dc2626;color:#dc2626}}
.pill[data-cat="Medium"].active{{border-color:#d97706;color:#d97706}}
.pill[data-cat="Small"].active{{border-color:#2563eb;color:#2563eb}}
.pill[data-cat="Wholesaler"].active{{border-color:#7c3aed;color:#7c3aed}}
.pill[data-status="Installed"].active{{border-color:var(--green);color:var(--green)}}
.pill[data-status="Open"].active{{border-color:var(--orange);color:var(--orange)}}
.pill[data-status="Potential"].active{{border-color:var(--accent);color:var(--accent)}}
.pill[data-owner="Thomas"].active{{border-color:var(--thomas);color:var(--thomas)}}
.pill[data-owner="Max"].active{{border-color:var(--max);color:var(--max)}}
.pill[data-country].active{{border-color:#6b7280;color:#374151}}
.pill:hover{{background:rgba(0,0,0,.04)}}

.stats{{display:flex;gap:14px;margin-left:auto;flex-shrink:0}}
.stat{{text-align:center;line-height:1.2}}
.stat-value{{font-size:16px;font-weight:700;color:var(--accent)}}
.stat-label{{font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}}

#map{{flex:1;width:100%}}

.leaflet-popup-content-wrapper{{
  background:var(--panel)!important;border:1.5px solid var(--border)!important;
  border-radius:var(--radius)!important;box-shadow:var(--shadow)!important;
  color:var(--text)!important;min-width:230px;
}}
.leaflet-popup-tip{{background:var(--panel)!important}}
.leaflet-popup-content{{margin:0!important;padding:0!important}}
.popup-inner{{padding:13px 15px}}
.popup-header{{border-bottom:1.5px solid var(--border);padding-bottom:8px;margin-bottom:8px}}
.popup-company{{font-size:13.5px;font-weight:700;color:var(--text);line-height:1.3;margin-bottom:3px}}
.popup-city-line{{font-size:10.5px;color:var(--accent);font-weight:600;letter-spacing:.03em}}
.popup-row{{
  display:flex;justify-content:space-between;align-items:center;
  font-size:11px;padding:3px 0;border-bottom:1px solid var(--border);
  color:var(--muted);gap:8px;
}}
.popup-row:last-child{{border-bottom:none}}
.popup-row strong{{color:var(--text);font-weight:600;text-align:right;max-width:65%;word-break:break-word}}
.badge{{display:inline-block;padding:1px 7px;border-radius:10px;font-size:10px;font-weight:700;color:#fff;white-space:nowrap}}
.owner-badge-thomas{{background:var(--thomas);color:#fff;font-size:10px;font-weight:700;padding:1px 7px;border-radius:10px}}
.owner-badge-max{{background:var(--max);color:#fff;font-size:10px;font-weight:700;padding:1px 7px;border-radius:10px}}

.leaflet-tooltip.company-label{{
  background:var(--panel);border:1.5px solid var(--border);border-radius:6px;
  color:var(--text);font-size:10.5px;font-weight:600;padding:2px 7px;
  box-shadow:var(--shadow);white-space:nowrap;pointer-events:none;
}}
.leaflet-tooltip.company-label::before{{display:none}}

#legend{{
  position:absolute;bottom:28px;right:10px;z-index:900;
  background:var(--panel);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:11px 13px;min-width:150px;box-shadow:var(--shadow);font-size:11px;
}}
#legend h3{{font-size:9.5px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:7px}}
.leg-item{{display:flex;align-items:center;gap:6px;margin-bottom:5px;color:var(--text)}}
.leg-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
.leg-diamond{{width:8px;height:8px;transform:rotate(45deg);flex-shrink:0}}
.leg-square{{width:8px;height:8px;border-radius:2px;flex-shrink:0}}
.leg-divider{{height:1px;background:var(--border);margin:7px 0}}

#search-wrap{{position:absolute;top:10px;left:50%;transform:translateX(-50%);z-index:900;width:260px}}
#search{{
  width:100%;padding:7px 13px;border-radius:24px;border:1.5px solid var(--border);
  background:var(--panel);color:var(--text);font-size:12px;outline:none;
  box-shadow:var(--shadow);transition:border-color .15s;
}}
#search:focus{{border-color:var(--accent)}}
#search::placeholder{{color:var(--muted)}}

#toast{{
  position:absolute;top:54px;left:50%;transform:translateX(-50%);
  background:var(--panel);border:1.5px solid var(--border);padding:6px 15px;
  border-radius:20px;font-size:11.5px;color:var(--muted);z-index:950;
  pointer-events:none;opacity:0;transition:opacity .3s;white-space:nowrap;
}}
#toast.show{{opacity:1}}

#label-toggle{{
  position:absolute;bottom:28px;left:10px;z-index:900;
  background:var(--panel);border:1.5px solid var(--border);border-radius:var(--radius);
  padding:7px 13px;font-size:11.5px;cursor:pointer;color:var(--muted);box-shadow:var(--shadow);transition:all .15s;
}}
#label-toggle:hover{{color:var(--text)}}
#label-toggle.on{{color:var(--accent);border-color:var(--accent)}}
</style>
</head>
<body>

<header>
  <h1>EU Installer Map <span>· {len(records)} installers</span></h1>
  <div class="pill-group">
    <span class="group-label">Owner</span>
    <div class="pill active" data-owner="Thomas" onclick="toggleOwner('Thomas')">
      <span class="dot" style="background:var(--thomas)"></span>Thomas</div>
    <div class="pill active" data-owner="Max" onclick="toggleOwner('Max')">
      <span class="dot" style="background:var(--max)"></span>Max</div>
    <div class="sep"></div>
    <span class="group-label">Category</span>
    <div class="pill active" data-cat="Large"      onclick="toggleCat('Large')">
      <span class="dot" style="background:#dc2626"></span>Large</div>
    <div class="pill active" data-cat="Medium"     onclick="toggleCat('Medium')">
      <span class="dot" style="background:#d97706"></span>Medium</div>
    <div class="pill active" data-cat="Small"      onclick="toggleCat('Small')">
      <span class="dot" style="background:#2563eb"></span>Small</div>
    <div class="pill active" data-cat="Wholesaler" onclick="toggleCat('Wholesaler')">
      <span class="dot" style="background:#7c3aed"></span>Wholesaler</div>
    <div class="sep"></div>
    <span class="group-label">Status</span>
    <div class="pill active" data-status="Installed"  onclick="toggleStatus('Installed')">✅ Installed</div>
    <div class="pill active" data-status="Open"       onclick="toggleStatus('Open')">🟠 Open</div>
    <div class="pill active" data-status="Potential"  onclick="toggleStatus('Potential')">🔵 Potential</div>
    <div class="sep"></div>
    <span class="group-label">Country</span>
    {country_pills_html}
  </div>
  <div class="stats">
    <div class="stat"><div class="stat-value" id="s-total">0</div><div class="stat-label">Total</div></div>
    <div class="stat"><div class="stat-value" id="s-thomas">0</div><div class="stat-label" style="color:var(--thomas)">Thomas</div></div>
    <div class="stat"><div class="stat-value" id="s-max">0</div><div class="stat-label" style="color:var(--max)">Max</div></div>
    <div class="stat"><div class="stat-value" id="s-installed">0</div><div class="stat-label">Installed</div></div>
  </div>
</header>

<div id="map"></div>
<div id="search-wrap">
  <input id="search" type="text" placeholder="🔍 Search company, city, contact…" oninput="doSearch(this.value)">
</div>
<button id="label-toggle" onclick="toggleLabels()">🏷 Labels: OFF</button>

<div id="legend">
  <h3>Owner (ring)</h3>
  <div class="leg-item"><span class="leg-dot" style="border:2.5px solid var(--thomas);background:transparent"></span>Thomas</div>
  <div class="leg-item"><span class="leg-dot" style="border:2.5px solid var(--max);background:transparent"></span>Max</div>
  <div class="leg-divider"></div>
  <h3>Category (fill)</h3>
  <div class="leg-item"><span class="leg-dot" style="background:#dc2626"></span>Large</div>
  <div class="leg-item"><span class="leg-dot" style="background:#d97706"></span>Medium</div>
  <div class="leg-item"><span class="leg-dot" style="background:#2563eb"></span>Small</div>
  <div class="leg-item"><span class="leg-dot" style="background:#7c3aed"></span>Wholesaler</div>
  <div class="leg-divider"></div>
  <h3>Shape = Country</h3>
  <div class="leg-item"><span class="leg-dot" style="background:#888"></span>Germany</div>
  <div class="leg-item"><span class="leg-diamond" style="background:#888"></span>Austria</div>
  <div class="leg-item"><span class="leg-square" style="background:#888"></span>Others</div>
  <div class="leg-divider"></div>
  <h3>Precision</h3>
  <div class="leg-item"><span class="leg-dot" style="background:#888;opacity:1"></span>Exact / ~Company</div>
  <div class="leg-item"><span class="leg-dot" style="background:#888;opacity:0.45;outline:1px dashed #888"></span>Country approx.</div>
</div>
<div id="toast"></div>

<script>
const RAW_DATA  = {geojson_data};
const SUMMARY   = {summary_json};

const CAT_COLOR  = {{"Large":"#dc2626","Medium":"#d97706","Small":"#2563eb","Wholesaler":"#7c3aed"}};
const CAT_RADIUS = {{"Large":12,"Medium":9,"Wholesaler":9,"Small":7}};
const OWNER_RING = {{"Thomas":"#0ea5e9","Max":"#f97316"}};
const STATUS_BADGE = {{
  "Installed":{{bg:"#dcfce7",color:"#15803d"}},
  "Open":{{bg:"#ffedd5",color:"#c2410c"}},
  "Potential":{{bg:"#dbeafe",color:"#1d4ed8"}},
}};

function makeIcon(cat, country, owner, precision) {{
  const fill = CAT_COLOR[cat]    || "#888";
  const ring = OWNER_RING[owner] || "#aaa";
  const r    = CAT_RADIUS[cat]   || 7;
  const size = (r + 4) * 2;
  const cx = size / 2, cy = size / 2;
  const fillOpacity = precision === "country" ? "0.4" : "1";
  const dash = precision === "country" ? 'stroke-dasharray="3,2"' : "";
  let shape;
  if (country === "Austria") {{
    const h = r + 2;
    shape = `<polygon points="${{cx}},${{cy-h}} ${{cx+h}},${{cy}} ${{cx}},${{cy+h}} ${{cx-h}},${{cy}}"
      fill="${{fill}}" fill-opacity="${{fillOpacity}}" stroke="${{ring}}" stroke-width="2.8" ${{dash}}/>`;
  }} else if (country === "Germany") {{
    shape = `<circle cx="${{cx}}" cy="${{cy}}" r="${{r}}"
      fill="${{fill}}" fill-opacity="${{fillOpacity}}" stroke="${{ring}}" stroke-width="2.8" ${{dash}}/>`;
  }} else {{
    shape = `<rect x="${{cx-r}}" y="${{cy-r}}" width="${{r*2}}" height="${{r*2}}" rx="3"
      fill="${{fill}}" fill-opacity="${{fillOpacity}}" stroke="${{ring}}" stroke-width="2.8" ${{dash}}/>`;
  }}
  return L.divIcon({{
    html:`<svg xmlns="http://www.w3.org/2000/svg" width="${{size}}" height="${{size}}">${{shape}}</svg>`,
    className:"",iconSize:[size,size],iconAnchor:[cx,cy],popupAnchor:[0,-cy]
  }});
}}

const map = L.map("map",{{center:[50.5,10.5],zoom:5}});
L.tileLayer("https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png",{{
  attribution:"© OpenStreetMap © CARTO",subdomains:"abcd",maxZoom:19
}}).addTo(map);

const clusterGroup = L.markerClusterGroup({{
  maxClusterRadius:45,
  iconCreateFunction(cluster) {{
    const n = cluster.getChildCount();
    return L.divIcon({{
      html:`<div style="background:rgba(37,99,235,.15);border:2px solid #2563eb;border-radius:50%;
        width:36px;height:36px;display:flex;align-items:center;justify-content:center;
        font-size:12px;font-weight:700;color:#1d4ed8">${{n}}</div>`,
      className:"",iconSize:[36,36],iconAnchor:[18,18]
    }});
  }}
}});

let allMarkers = [];
let activeCats      = new Set(["Large","Medium","Small","Wholesaler"]);
let activeStatuses  = new Set(["Installed","Open","Potential","Unknown","Private","active","Active"]);
let activeOwners    = new Set(["Thomas","Max"]);
let activeCountries = new Set({json.dumps(all_countries)});
let showLabels = false;

function popupHTML(p) {{
  const catColor  = CAT_COLOR[p.category] || "#888";
  const sb        = STATUS_BADGE[p.status] || {{bg:"#f3f4f6",color:"#6b7280"}};
  const ownerBadge = p.owner === "Thomas"
    ? `<span class="owner-badge-thomas">Thomas</span>`
    : `<span class="owner-badge-max">Max</span>`;
  const actLine     = p.act     ? `<div class="popup-row"><span>Act</span><strong>${{p.act}}</strong></div>` : "";
  const invLine     = p.inverter && p.inverter !== "—" ? `<div class="popup-row"><span>Inverter</span><strong>${{p.inverter}}</strong></div>` : "";
  const contactLine = p.contact && p.contact !== "—"  ? `<div class="popup-row"><span>Contact</span><strong>${{p.contact}}</strong></div>` : "";
  const phoneLine   = p.phone   && p.phone !== "—"    ? `<div class="popup-row"><span>Phone</span><strong>${{p.phone}}</strong></div>` : "";
  const precNote    = p.precision === "country"
    ? `<div style="margin-top:6px;padding:4px 8px;background:#fef9c3;border:1px solid #fde047;border-radius:6px;font-size:10px;color:#854d0e">⚠ Approximate location (no address – showing country centre)</div>`
    : p.precision === "approximate"
    ? `<div style="margin-top:6px;padding:4px 8px;background:#f0fdf4;border:1px solid #86efac;border-radius:6px;font-size:10px;color:#166534">~ Located by company name</div>`
    : "";
  return `
  <div class="popup-inner">
    <div class="popup-header">
      <div class="popup-company">${{p.company}}</div>
      <div class="popup-city-line">📍 ${{p.city || p.country}}, ${{p.country}} &nbsp;${{ownerBadge}}</div>
    </div>
    <div class="popup-row"><span>Category</span>
      <strong><span class="badge" style="background:${{catColor}}">${{p.category}}</span></strong></div>
    <div class="popup-row"><span>Status</span>
      <strong><span class="badge" style="background:${{sb.bg}};color:${{sb.color}};border:1px solid ${{sb.color}}">${{p.status}}</span></strong></div>
    ${{actLine}}${{invLine}}${{contactLine}}${{phoneLine}}${{precNote}}
  </div>`;
}}

for (const f of RAW_DATA.features) {{
  const p = f.properties;
  const [lon, lat] = f.geometry.coordinates;
  const marker = L.marker([lat,lon],{{icon:makeIcon(p.category,p.country,p.owner,p.precision)}});
  marker.bindPopup(popupHTML(p),{{maxWidth:300}});
  marker.bindTooltip(`${{p.company}} (${{p.owner}})`,{{
    permanent:false,direction:"top",className:"company-label",offset:[0,-4]
  }});
  allMarkers.push({{marker,props:p}});
}}
clusterGroup.addTo(map);
renderMarkers();

function renderMarkers() {{
  clusterGroup.clearLayers();
  let tot=0,thomas=0,max=0,inst=0;
  for (const {{marker,props}} of allMarkers) {{
    if (activeCats.has(props.category) &&
        activeStatuses.has(props.status) &&
        activeOwners.has(props.owner) &&
        activeCountries.has(props.country)) {{
      clusterGroup.addLayer(marker);
      tot++;
      if (props.owner==="Thomas") thomas++;
      if (props.owner==="Max")    max++;
      if (props.status==="Installed") inst++;
    }}
  }}
  document.getElementById("s-total").textContent    = tot;
  document.getElementById("s-thomas").textContent   = thomas;
  document.getElementById("s-max").textContent      = max;
  document.getElementById("s-installed").textContent= inst;
  if (showLabels) applyLabels(true);
}}

function toggleOwner(o)   {{ activeOwners.has(o)    ?activeOwners.delete(o)   :activeOwners.add(o);    document.querySelector(`[data-owner="${{o}}"]`).classList.toggle("active");   renderMarkers(); }}
function toggleCat(cat)   {{ activeCats.has(cat)     ?activeCats.delete(cat)   :activeCats.add(cat);    document.querySelector(`[data-cat="${{cat}}"]`).classList.toggle("active");     renderMarkers(); }}
function toggleStatus(s)  {{ activeStatuses.has(s)   ?activeStatuses.delete(s) :activeStatuses.add(s);  document.querySelector(`[data-status="${{s}}"]`).classList.toggle("active");  renderMarkers(); }}
function toggleCountry(c) {{ activeCountries.has(c)  ?activeCountries.delete(c):activeCountries.add(c); document.querySelector(`[data-country="${{c}}"]`).classList.toggle("active"); renderMarkers(); }}

function applyLabels(on) {{
  for (const {{marker,props}} of allMarkers) {{
    marker.unbindTooltip();
    marker.bindTooltip(`${{props.company}} (${{props.owner}})`,{{
      permanent:on,direction:"top",className:"company-label",offset:[0,-4]
    }});
  }}
}}
function toggleLabels() {{
  showLabels=!showLabels;
  applyLabels(showLabels);
  const btn=document.getElementById("label-toggle");
  btn.textContent=`🏷 Labels: ${{showLabels?"ON":"OFF"}}`;
  btn.classList.toggle("on",showLabels);
}}

let searchTimer;
function doSearch(val) {{
  clearTimeout(searchTimer);
  searchTimer=setTimeout(()=>{{
    const q=val.trim().toLowerCase();
    if(!q) return;
    const hit=allMarkers.find(m=>
      m.props.company.toLowerCase().includes(q)||
      m.props.city.toLowerCase().includes(q)||
      m.props.contact.toLowerCase().includes(q)
    );
    if(hit){{ map.flyTo(hit.marker.getLatLng(),12,{{duration:1.2}}); setTimeout(()=>hit.marker.openPopup(),1300); }}
    else showToast("Not found: "+val);
  }},350);
}}
function showToast(msg) {{
  const el=document.getElementById("toast");
  el.textContent=msg; el.classList.add("show");
  setTimeout(()=>el.classList.remove("show"),2500);
}}
</script>
</body>
</html>"""

out = Path(OUTPUT_HTML)
out.write_text(HTML, encoding="utf-8")
print(f"\n✅ 地图已生成: {out.resolve()}")
print("   👉 双击 HTML 用浏览器打开，或上传 Netlify Drop 分享")

import os, sys, subprocess
if sys.platform == "win32":
    os.startfile(str(out.resolve()))
elif sys.platform == "darwin":
    subprocess.run(["open", str(out.resolve())])
else:
    subprocess.run(["xdg-open", str(out.resolve())])